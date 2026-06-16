import os
import re
import sys
import glob

import openpyxl
import pytest

# Add the script directory to the path so we can import the generator.
sys.path.append(os.path.join(os.path.dirname(__file__), '../terraform/scripts'))
import generate_delivery as gd  # noqa: E402

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
ORG_MAIN_TF = os.path.join(REPO_ROOT, 'terraform', '2_organization', 'main.tf')

RESOURCES_HEADERS = [
    "resource_type", "parent_name", "resource_name", "environment",
    "existing_project_id", "owner", "budget_amount", "budget_alert_emails",
    "shared_vpc", "vpc_sc", "central_monitoring", "central_logging", "org_tags",
]


def _write_ssot(path):
    """テスト用の最小 SSoT(xlsx) を作成する（フォルダ1・プロジェクト1・各シート最小限）。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("resources")
    ws.append(RESOURCES_HEADERS)
    ws.append(["folder", "organization_id", "admin", "", "", "", "", "", "", "", "", "", ""])
    ws.append([
        "project", "admin", "my-app", "dev", "", "team-x",
        "10000", "a@example.com", "", "", "TRUE", "TRUE", "cost-center/x",
    ])

    ws = wb.create_sheet("org_policies")
    ws.append(["target_name", "policy_id", "enforce", "allow_list", "apply_mode"])
    ws.append(["organization_id", "iam.disableServiceAccountKeyCreation", "TRUE", "", "live"])

    ws = wb.create_sheet("log_sinks")
    ws.append(["log_type", "filter", "destination_type", "destination_parent", "retention_days"])
    ws.append(["admin-activity", 'logName:"activity"', "BigQuery", "audit_logs", "365"])

    ws = wb.create_sheet("tag_definitions")
    ws.append(["tag_key", "allowed_values", "description"])
    ws.append(["cost-center", "x,y", "コストセンター"])

    # shared_vpc_subnets / vpc_sc_perimeters はあえて空（=ネットワーク項目は「—」になることを検証）
    wb.create_sheet("shared_vpc_subnets").append(
        ["host_project_env", "subnet_name", "region", "ip_cidr_range"])
    wb.create_sheet("vpc_sc_perimeters").append(
        ["perimeter_name", "title", "restricted_services", "dry_run"])

    wb.save(path)


def _write_common_tfvars(path, simplified=True, group_iam=False):
    path.write_text(
        'organization_domain = "example.com"\n'
        'project_id_prefix   = "ex"\n'
        'billing_account_id  = "AAAA-BBBB-CCCC"\n'
        f'enable_simplified_admin_groups = {str(simplified).lower()}\n'
        f'enable_group_iam               = {str(group_iam).lower()}\n'
        'enable_tags = true\n',
        encoding="utf-8",
    )


@pytest.fixture
def generated(tmp_path, monkeypatch):
    """最小 SSoT から納品物を生成し、開いた Workbook と全セル文字列を返す。"""
    xlsx = tmp_path / "gcp-foundations.xlsx"
    tfvars = tmp_path / "common.tfvars"
    out_dir = tmp_path / "delivery"
    _write_ssot(str(xlsx))
    _write_common_tfvars(tfvars, simplified=True, group_iam=False)

    monkeypatch.setattr(gd, "XLSX_PATH", str(xlsx))
    monkeypatch.setattr(gd, "COMMON_TFVARS", str(tfvars))
    monkeypatch.setattr(gd, "DOMAIN_ENV", str(tmp_path / "domain.env"))  # 存在しない＝フォールバック
    monkeypatch.setattr(gd, "OUTPUT_DIR", str(out_dir))
    monkeypatch.setattr(gd, "ROOT_DIR", str(tmp_path))
    # 表紙メタが環境変数に影響されないよう既定化
    for k in ("DELIVERY_CUSTOMER", "DELIVERY_AUTHOR", "DELIVERY_VERSION", "DELIVERY_DOCNO"):
        monkeypatch.delenv(k, raising=False)

    gd.main()

    files = glob.glob(os.path.join(str(out_dir), "*.xlsx"))
    assert len(files) == 1, f"想定どおり1ファイルが生成されること: {files}"
    wb = openpyxl.load_workbook(files[0])
    return wb


def _all_text(ws):
    out = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                out.append(str(v))
    return "\n".join(out)


# --- 生成・シート構成 -----------------------------------------------------------

def test_all_expected_sheets_present(generated):
    expected = {
        "表紙", "改訂履歴", "目次",
        "1.構築項目一覧", "2.構築概要", "3.フォルダ構成", "4.プロジェクト一覧",
        "5.組織ポリシー", "6.ログ集約シンク", "7.BigQuery リソース", "8.監視・予算",
        "9.ネットワーク", "10.タグ・ラベル", "11.Googleグループ・IAM", "12.費用注意事項",
    }
    assert expected.issubset(set(generated.sheetnames))


def test_bigquery_sheet_lists_log_sink_and_asset_inventory(generated):
    text = _all_text(generated["7.BigQuery リソース"])
    assert "audit_logs" in text          # log_sinks の BigQuery 宛先データセット
    assert "asset_inventory" in text     # IAM 監査の基盤標準データセット
    assert "v_iam_policy" in text


# --- 構築項目一覧の実施状況（SSoT 定義有無による 完了/—）-------------------------

def test_summary_status_reflects_ssot(generated):
    text = _all_text(generated["1.構築項目一覧"])
    # フォルダ定義あり → 完了 / サブネット・境界なし → ネットワークは「—」
    assert "管理フォルダ構成" in text
    assert "組織ポリシー設定" in text
    assert "—" in text  # ネットワーク構成が未定義なので「—」が存在する


def test_network_marked_not_applicable(generated):
    rows = list(generated["1.構築項目一覧"].iter_rows(values_only=True))
    net = [r for r in rows if r and r[1] == "ネットワーク構成"]
    assert net, "ネットワーク構成 行が存在すること"
    assert net[0][3] == "—", "サブネット/境界が無いので実施状況は — であること"


# --- ラベル（app は常に、env/owner は値がある時のみ）-----------------------------

def test_labels_rendered(generated):
    text = _all_text(generated["10.タグ・ラベル"])
    assert "app=my-app" in text
    assert "env=dev" in text
    assert "owner=team-x" in text


# --- グループ／IAM（集約モード・IAM 未適用が反映される）-------------------------

def test_groups_simplified_and_iam_off(generated):
    text = _all_text(generated["11.Googleグループ・IAM"])
    assert "集約モード（2グループ）" in text
    assert "未適用" in text
    assert "gcp-organization-admins@example.com" in text
    assert "gcp-billing-admins@example.com" in text


# --- resolve_group_roles のロジック ---------------------------------------------

def test_resolve_group_roles_full_has_nine_groups():
    groups = gd.resolve_group_roles(simplified=False)
    assert len(groups) == 9


def test_resolve_group_roles_simplified_merges_non_billing():
    groups = gd.resolve_group_roles(simplified=True)
    assert len(groups) == 2
    by_name = {g: roles for g, _desc, roles in groups}
    # 重複なし
    assert len(by_name["gcp-organization-admins"]) == len(set(by_name["gcp-organization-admins"]))
    # 請求ロールは org-admins に含まれない（billing は別グループ）
    assert "roles/billing.creator" not in by_name["gcp-organization-admins"]
    assert "roles/billing.creator" in by_name["gcp-billing-admins"]
    # 非請求の代表ロールが org-admins に統合されている
    assert "roles/resourcemanager.organizationAdmin" in by_name["gcp-organization-admins"]


# --- ドリフト検知: GROUP_ROLES が terraform の raw_roles と一致すること -----------

def _parse_raw_roles_from_terraform():
    with open(ORG_MAIN_TF, "r", encoding="utf-8") as f:
        text = f.read()
    start = text.index("raw_roles = {")
    end = text.index("simplified_group_roles")
    block = text[start:end]
    result = {}
    for m in re.finditer(r'"(gcp-[a-z-]+)"\s*=\s*\[(.*?)\]', block, re.S):
        name = m.group(1)
        roles = re.findall(r'(roles/[A-Za-z0-9.]+)', m.group(2))
        result[name] = set(roles)
    return result


def test_group_roles_match_terraform_no_drift():
    tf = _parse_raw_roles_from_terraform()
    py = {g: set(roles) for g, (_desc, roles) in gd.GROUP_ROLES.items()}
    assert tf, "terraform 側の raw_roles が解析できること"
    assert set(tf.keys()) == set(py.keys()), (
        "グループ集合が terraform と一致すること。"
        "terraform/2_organization/main.tf を変更したら GROUP_ROLES も更新してください。"
    )
    for g in tf:
        assert tf[g] == py[g], (
            f"グループ '{g}' のロールが terraform と不一致。"
            "generate_delivery.py の GROUP_ROLES を更新してください。"
        )
