# /// script
# requires-python = ">=3.12"
# dependencies = [
#     "openpyxl==3.1.2",
# ]
# ///
import os
import shutil
import sys

try:
    import openpyxl
except Exception:
    print("❌ openpyxl not found. Run 'make install' first.")
    sys.exit(1)


def prune_orphans():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    projects_dir = os.path.normpath(os.path.join(script_dir, '..', '4_projects'))
    xlsx_path = os.path.normpath(os.path.join(script_dir, '..', '..', 'gcp-foundations.xlsx'))

    valid_project_names: set[str] = set()
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        if 'resources' in wb.sheetnames:
            ws = wb['resources']
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = dict(zip(headers, row))
                res_type = str(row_dict.get('resource_type') or '').strip().lower()
                res_name = str(row_dict.get('resource_name') or '').strip()
                if res_type == 'project' and res_name:
                    valid_project_names.add(res_name)
    else:
        print("ℹ️  gcp-foundations.xlsx not found. Treating all project directories as orphans.")

    if not os.path.exists(projects_dir):
        print("✅ terraform/4_projects/ directory does not exist. Nothing to prune.")
        return

    orphans: list[str] = []
    for entry in sorted(os.listdir(projects_dir)):
        if entry == 'template' or entry.startswith('.'):
            continue
        proj_path = os.path.join(projects_dir, entry)
        if os.path.isdir(proj_path) and entry not in valid_project_names:
            orphans.append(entry)

    if not orphans:
        print("✅ No orphan directories found. terraform/4_projects/ is in sync with SSoT.")
        return

    print(f"Found {len(orphans)} orphan project director{'y' if len(orphans) == 1 else 'ies'} not defined in gcp-foundations.xlsx:\n")
    for name in orphans:
        tfvars = os.path.join(projects_dir, name, 'terraform.tfvars')
        status = "(terraform.tfvars present — GCP resources may still exist!)" if os.path.exists(tfvars) else "(already excluded from deployment)"
        print(f"  - terraform/4_projects/{name}/  {status}")

    print()
    print("⚠️  IMPORTANT: Before pruning, verify that GCP resources in these projects have already been destroyed.")
    print("   If unsure, check the GCP console or run 'terraform state list' in each directory.")
    print()
    answer = input("Delete these directories? (type 'PRUNE' to confirm, anything else to abort): ").strip()
    if answer != 'PRUNE':
        print("Aborted.")
        sys.exit(0)

    print()
    for name in orphans:
        proj_path = os.path.join(projects_dir, name)
        shutil.rmtree(proj_path)
        print(f"  🗑️  Deleted terraform/4_projects/{name}/")

    print()
    print("✅ Prune complete.")


if __name__ == "__main__":
    prune_orphans()
