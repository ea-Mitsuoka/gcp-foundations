# /// script
# requires-python = ">=3.12"
# dependencies = [
#     "openpyxl==3.1.2",
# ]
# ///
import os
import shutil
import glob
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception:
    openpyxl = None
    class Workbook:
        def __init__(self):
            self.active = None
    DataValidation = None
import json
import sys
import re
import ipaddress
import csv

def to_snake_case(name):
    name = re.sub(r'[\.\s-]', '_', str(name))
    name = re.sub(r'(.)([A-Z][a-z]+)', r'\1_\2', name)
    name = re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', name).lower()
    return name

# 環境(environment)の決定ロジック。
# 方針: `environment` 列は任意。名前からの推定は一切しない（名前と環境を分離）。
#       指定時は prod/stag/dev に限定。空欄は ''（=env ラベルを付けない）。
ENV_ALLOWED = ("prod", "stag", "dev")

def resolve_environment(explicit_env):
    """環境を決定する。返り値 (env, error)。
    任意項目。空欄は '' を返し env ラベルを付与しない（暗黙の推定・デフォルトはしない）。
    指定時は prod/stag/dev に限定し、許可外はエラー。
    """
    explicit = str(explicit_env or "").strip().lower()
    if not explicit:
        return ("", None)
    if explicit not in ENV_ALLOWED:
        return (None, f"environment '{explicit}' は許可外です（{', '.join(ENV_ALLOWED)} のいずれか、または空欄）")
    return (explicit, None)

def resolve_shared_vpc_env(env, shared_vpc):
    """接続先 Shared VPC ホストの環境を決定する。ホストは prod/dev の2つのみ（stag は prod ホストに相乗り）。
    shared_vpc 未指定なら 'none'。environment から導出し、命名接頭辞には依存しない。"""
    if not str(shared_vpc or "").strip():
        return "none"
    return "dev" if env == "dev" else "prod"

class ResourceValidator:
    @staticmethod
    def validate_gcp_resource_name(name, resource_type, prefix=""):
        if not name: return "Resource name is empty."
        name = str(name).strip()
        if resource_type == 'project':
            full_name = f"{prefix}-{name}" if prefix and prefix != "unknown" else name
            if not re.match(r'^[a-z][a-z0-9-]{4,28}[a-z0-9]$', full_name):
                if len(full_name) > 30:
                    return f"Project ID '{full_name}' exceeds the 30-character limit (prefix '{prefix}' + name '{name}')."
                return f"Project ID '{full_name}' is invalid. Must be 6-30 chars, lowercase letters, numbers, and hyphens."
        elif resource_type == 'folder':
            if not re.match(r'^[a-zA-Z0-9- ]{1,30}$', name):
                return f"Folder name '{name}' is invalid."
        return None

    @staticmethod
    def validate_cidr(cidr, used_cidrs):
        if not cidr: return None
        try:
            new_net = ipaddress.ip_network(cidr, strict=True)
            for old_net in used_cidrs:
                if new_net.overlaps(old_net):
                    return f"CIDR '{cidr}' overlaps with '{old_net}'."
            return None
        except ValueError as e:
            return f"Invalid CIDR format: '{cidr}' ({e})."

    @staticmethod
    def validate_hierarchy(resources):
        errors = []
        seen_names = set()
        folders = {str(r['resource_name']).strip() for r in resources if str(r['resource_type']).strip().lower() == 'folder'}
        for r in resources:
            name = str(r['resource_name']).strip()
            parent = str(r['parent_name']).strip()
            res_type = str(r['resource_type']).strip().lower()
            
            if name in seen_names:
                errors.append(f"Duplicate resource name '{name}' detected. Names must be unique across all folders and projects.")
            seen_names.add(name)

            if name == parent:
                errors.append(f"Resource '{name}' circular reference.")
            if parent != 'organization_id' and parent not in folders:
                errors.append(f"{res_type.capitalize()} '{name}' refers to parent '{parent}' which is not defined.")
        return errors

    @staticmethod
    def validate_tags(org_tags_str, tag_definitions):
        if not org_tags_str: return None
        tags = [t.strip() for t in org_tags_str.split(',') if t.strip()]
        for tag in tags:
            if '/' not in tag: return f"Invalid tag format '{tag}'."
            key, val = tag.split('/', 1)
            if key not in tag_definitions: return f"Tag key '{key}' undefined."
            if val not in tag_definitions[key]['allowed_values']: return f"Tag value '{val}' not allowed for '{key}'."
        return None

    @staticmethod
    def validate_alerts(notifications, alert_defs):
        errors = []
        seen_alerts = set()
        
        for idx, alert in enumerate(alert_defs, start=2):
            name = str(alert.get('alert_name') or '').strip()
            if name:
                if name in seen_alerts:
                    errors.append(f"[alert_definitions] Row {idx}: Duplicate alert_name '{name}'. Alert names must be unique.")
                seen_alerts.add(name)
            
            # --- ▼ 追加: alert_documentation の必須バリデーション ---
            doc = str(alert.get('alert_documentation') or '').strip()
            if not doc:
                errors.append(f"[alert_definitions] Row {idx}: 'alert_documentation' (説明) は必須項目です。空欄は許可されません。")

        for idx, n in enumerate(notifications, start=2):
            name = str(n.get('alert_name') or '').strip()
            if name and name not in seen_alerts:
                errors.append(f"[notifications] Row {idx}: Refers to undefined alert '{name}'.")
        return errors

    @staticmethod
    def validate_project_refs(resources, subnets, perimeters):
        errors = []
        subnet_names = {str(s.get('subnet_name') or '').strip() for s in subnets if s.get('subnet_name')}
        perimeter_names = {str(p.get('perimeter_name') or '').strip() for p in perimeters if p.get('perimeter_name')}
        for idx, r in enumerate(resources, start=2):
            if str(r.get('resource_type') or '').strip().lower() == 'folder': continue
            subnet = str(r.get('shared_vpc') or '').strip()
            if subnet and subnet not in subnet_names:
                errors.append(f"[resources] Row {idx}: Refers to undefined shared_vpc subnet '{subnet}'.")
            vpc_sc = str(r.get('vpc_sc') or '').strip()
            if vpc_sc and vpc_sc not in perimeter_names:
                errors.append(f"[resources] Row {idx}: Refers to undefined vpc_sc perimeter '{vpc_sc}'.")
        return errors

    @staticmethod
    def validate_org_policies(org_policies, folders, projects):
        errors = []
        valid_targets = {'organization_id'} | folders | projects
        valid_modes = {'live', 'dryrun', 'both'}
        for idx, p in enumerate(org_policies, start=2):
            target = str(p.get('target_name') or '').strip()
            if target and target not in valid_targets:
                errors.append(f"[org_policies] Row {idx}: Target '{target}' is not defined as a folder or project.")
            # apply_mode: live(spec のみ) / dryrun(dry_run_spec のみ) / both(両方)。空欄は live 扱い（後方互換）。
            mode = str(p.get('apply_mode') or '').strip().lower()
            if mode and mode not in valid_modes:
                errors.append(f"[org_policies] Row {idx}: apply_mode '{mode}' は不正です。'live' / 'dryrun' / 'both' のいずれか、または空欄（=live）を指定してください。")
        return errors

    @staticmethod
    def validate_log_sinks(log_sinks):
        errors = []
        seen_types = set()
        for idx, sink in enumerate(log_sinks, start=2):
            l_type = str(sink.get('log_type') or '').strip()
            if l_type:
                if l_type in seen_types:
                    errors.append(f"[log_sinks] Row {idx}: Duplicate log_type '{l_type}'. Log types must be unique.")
                seen_types.add(l_type)
            
            # --- ▼ 追加: BigQuery宛先のハイフン禁止バリデーション ---
            dest_type = str(sink.get('destination_type') or '').strip().lower()
            dest_parent = str(sink.get('destination_parent') or '').strip()
            if dest_type == 'bigquery' and '-' in dest_parent:
                errors.append(f"[log_sinks] Row {idx}: BigQueryのデータセット名 '{dest_parent}' にハイフン(-)は使用できません。アンダースコア(_)を使用してください。")
        return errors

def sanitize_id(name):
    if not name: return "unknown"
    return str(name).replace("-", "_").replace(" ", "_").replace(".", "_")

def generate_resources():
    # --- Cleanup Phase ---
    print("🧹 Cleaning up previously auto-generated files...")
    terraform_dir = os.path.join(os.path.dirname(__file__), '..')
    for file_path in glob.glob(os.path.join(terraform_dir, '**', 'auto_*.tf'), recursive=True):
        print(f"  - Deleting {file_path}")
        os.remove(file_path)

    # --- Setup Phase ---
    domain_env_path = os.path.join(os.path.dirname(__file__), '../../domain.env')
    domain = ""
    if os.path.exists(domain_env_path):
        with open(domain_env_path, 'r') as f:
            for line in f:
                if line.startswith('domain='):
                    domain = line.split('=')[1].strip().strip('"')
                    break

    project_id_prefix = "unknown"
    enable_shared_vpc = None
    enable_vpc_host_projects = None
    common_vars_path = os.path.join(os.path.dirname(__file__), '../common.tfvars')
    if os.path.exists(common_vars_path):
        with open(common_vars_path, 'r') as f:
            for line in f:
                line = line.strip()
                if line.startswith('project_id_prefix'):
                    project_id_prefix = line.split('=')[1].strip().strip('"').strip()
                elif line.startswith('enable_shared_vpc'):
                    enable_shared_vpc = line.split('=')[1].strip().strip('"').lower() == 'true'
                elif line.startswith('enable_vpc_host_projects'):
                    enable_vpc_host_projects = line.split('=')[1].strip().strip('"').lower() == 'true'
    mgmt_project_id = f"{project_id_prefix}-monitoring" 

    xlsx_path = os.path.join(os.path.dirname(__file__), '../../gcp-foundations.xlsx')
    if not os.path.exists(xlsx_path):
        print(f"'{xlsx_path}' not found. Creating a template...")
        wb = Workbook()
        ws = wb.active
        ws.title = "resources"
        headers = ["resource_type", "parent_name", "resource_name", "environment", "existing_project_id", "owner", "budget_amount", "budget_alert_emails", "shared_vpc", "vpc_sc", "central_monitoring", "central_logging", "org_tags", "billing_account"]
        ws.append(headers)
        wb.save(xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    required_sheets = {
        "resources": ["resource_type", "parent_name", "resource_name", "environment", "existing_project_id", "owner", "budget_amount", "budget_alert_emails", "shared_vpc", "vpc_sc", "central_monitoring", "central_logging", "org_tags", "billing_account"],
        "tag_definitions": ["tag_key", "allowed_values", "description"],
        "vpc_sc_perimeters": ["perimeter_name", "title", "restricted_services", "dry_run"],
        "vpc_sc_access_levels": ["access_level_name", "ip_subnetworks", "members"],
        "shared_vpc_subnets": ["host_project_env", "subnet_name", "region", "ip_cidr_range"],
        "org_policies": ["target_name", "policy_id", "enforce", "allow_list", "apply_mode"],
        "notifications": ["alert_name", "user_email", "receive_alerts"],
        "alert_definitions": ["alert_name", "alert_display_name", "metric_filter", "alert_documentation"],
        "log_sinks": ["log_type", "filter", "destination_type", "destination_parent", "retention_days"]
    }

    updated = False
    for sname, headers in required_sheets.items():
        if sname not in wb.sheetnames:
            ws = wb.create_sheet(sname)
            ws.append(headers)
            updated = True
    if updated: wb.save(xlsx_path)

    errors = []
    validator = ResourceValidator()
    
    tag_definitions = {}
    if 'tag_definitions' in wb.sheetnames:
        ws = wb['tag_definitions']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            t_dict = dict(zip(headers, row))
            key = str(t_dict.get('tag_key') or '').strip()
            if key:
                allowed = [v.strip() for v in str(t_dict.get('allowed_values') or '').split(',') if v.strip()]
                tag_definitions[key] = {'allowed_values': allowed, 'description': t_dict.get('description')}

    perimeters = []
    if 'vpc_sc_perimeters' in wb.sheetnames:
        ws = wb['vpc_sc_perimeters']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            perimeters.append(dict(zip(headers, row)))

    folders_map = {}
    projects = []
    resources_data = []
    if 'resources' in wb.sheetnames:
        ws = wb['resources']
        headers = [cell.value for cell in ws[1]]
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            row_dict = dict(zip(headers, row))
            resources_data.append(row_dict)
            res_name = str(row_dict.get('resource_name') or '').strip()
            res_type = str(row_dict.get('resource_type') or '').strip().lower()
            if res_type == 'folder': folders_map[res_name] = str(row_dict.get('parent_name') or '').strip()
            elif res_type == 'project': projects.append(row_dict)

    subnets_data = []
    if 'shared_vpc_subnets' in wb.sheetnames:
        ws = wb['shared_vpc_subnets']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row): subnets_data.append(dict(zip(headers, row)))

    log_sinks_data = []
    if 'log_sinks' in wb.sheetnames:
        ws = wb['log_sinks']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row): log_sinks_data.append(dict(zip(headers, row)))

    org_policies_data = []
    if 'org_policies' in wb.sheetnames:
        ws = wb['org_policies']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row): org_policies_data.append(dict(zip(headers, row)))

    alert_defs = []
    notifications = []
    if 'alert_definitions' in wb.sheetnames:
        ws = wb['alert_definitions']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row): alert_defs.append(dict(zip(headers, row)))
    if 'notifications' in wb.sheetnames:
        ws = wb['notifications']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row): notifications.append(dict(zip(headers, row)))

    if alert_defs:
        alert_errs = validator.validate_alerts(notifications, alert_defs)
        if alert_errs: errors.extend(alert_errs)
        
    if log_sinks_data:
        sink_errs = validator.validate_log_sinks(log_sinks_data)
        if sink_errs: errors.extend(sink_errs)

    if resources_data:
        hierarchy_errors = validator.validate_hierarchy(resources_data)
        if hierarchy_errors: errors.extend(hierarchy_errors)
        
        ref_errs = validator.validate_project_refs(resources_data, subnets_data, perimeters)
        if ref_errs: errors.extend(ref_errs)

        adopt_ids = {}
        for idx, r in enumerate(resources_data, start=2):
            existing_pid = str(r.get('existing_project_id') or '').strip()
            if existing_pid and str(r.get('resource_type') or '').strip().lower() == 'project':
                # 採用(adopt)モード: 実IDは existing_project_id。合成ID長の検証はスキップし、ID形式と重複のみ検証。
                if not re.match(r'^[a-z][a-z0-9-]{4,28}[a-z0-9]$', existing_pid):
                    errors.append(f"[resources] Row {idx}: existing_project_id '{existing_pid}' はGCPプロジェクトID形式（6-30文字・先頭英字・末尾英数字・小文字/数字/ハイフン）として不正です。")
                if existing_pid in adopt_ids:
                    errors.append(f"[resources] Row {idx}: existing_project_id '{existing_pid}' が Row {adopt_ids[existing_pid]} と重複しています（二重定義禁止）。")
                else:
                    adopt_ids[existing_pid] = idx
                rn = str(r.get('resource_name') or '').strip()
                if not re.match(r'^[a-z0-9]([a-z0-9-]{0,61}[a-z0-9])?$', rn):
                    errors.append(f"[resources] Row {idx}: resource_name '{rn}' は小文字英数字とハイフンのみにしてください（採用モードでも表示名/ラベル/ディレクトリ名に使用）。")
            else:
                name_err = validator.validate_gcp_resource_name(r.get('resource_name'), r.get('resource_type'), prefix=project_id_prefix)
                if name_err: errors.append(f"[resources] Row {idx}: {name_err}")
            # environment 列の検証（プロジェクトのみ）: 明示値の許可外・命名接頭辞との矛盾をエラーにする
            if str(r.get('resource_type') or '').strip().lower() == 'project':
                # 表示名(name)は resource_name をそのまま使う（環境サフィックスなし）ため、
                # GCP プロジェクト表示名の制約 4〜30 文字を満たすこと。
                rn_disp = str(r.get('resource_name') or '').strip()
                if not (4 <= len(rn_disp) <= 30):
                    errors.append(f"[resources] Row {idx}: resource_name '{rn_disp}' は4〜30文字にしてください（プロジェクト表示名として使用されるため）。")
                env_resolved, env_err = resolve_environment(r.get('environment'))
                if env_err:
                    errors.append(f"[resources] Row {idx}: {env_err}")
                # shared_vpc を使う場合は接続先ホスト(prod/dev)判定のため environment が必須
                elif str(r.get('shared_vpc') or '').strip() and not env_resolved:
                    errors.append(f"[resources] Row {idx}: shared_vpc を使う場合は environment（接続先ホスト判定用）の指定が必須です。")
            tag_err = validator.validate_tags(str(r.get('org_tags') or ''), tag_definitions)
            if tag_err: errors.append(f"[resources] Row {idx}: {tag_err}")
            
            # --- owner の GCPラベル形式バリデーション（任意。空欄=owner ラベルなし） ---
            owner = str(r.get('owner') or '').strip()
            if owner and not re.match(r'^[a-z0-9_-]{1,63}$', owner):
                errors.append(f"[resources] Row {idx}: Owner '{owner}' はGCPラベルとして不正な形式です。1-63文字の小文字、数字、ハイフン(-)、アンダースコア(_)のみを使用してください（@や.は不可）。空欄も可（その場合 owner ラベルは付与されません）。")

    if org_policies_data:
        folder_names = set(folders_map.keys())
        project_names = {p['resource_name'] for p in projects}
        op_errs = validator.validate_org_policies(org_policies_data, folder_names, project_names)
        if op_errs: errors.extend(op_errs)

    if subnets_data:
        used_cidrs = []
        for idx, s in enumerate(subnets_data, start=2):
            cidr = str(s.get('ip_cidr_range') or '').strip()
            if cidr:
                err = validator.validate_cidr(cidr, used_cidrs)
                if err: errors.append(f"[shared_vpc_subnets] Row {idx}: {err}")
                else: used_cidrs.append(ipaddress.ip_network(cidr, strict=True))

    if enable_shared_vpc is True and enable_vpc_host_projects is False:
        errors.append(
            "[common.tfvars] enable_shared_vpc = true には enable_vpc_host_projects = true が必要です。"
            " ホストプロジェクトが存在しないためサービスプロジェクトの接続がスキップされます。"
        )

    if errors:
        print("\n❌ Configuration errors detected:")
        for err in errors: print(f"  - {err}")
        sys.exit(1)


    # === 孤立プロジェクト (Orphan Projects) のクリーンアップ ===
    valid_project_names = {str(p.get('resource_name') or '').strip() for p in projects if p.get('resource_name')}
    projects_dir = os.path.join(os.path.dirname(__file__), '../4_projects')
    if os.path.exists(projects_dir):
        for entry in os.listdir(projects_dir):
            if entry == 'template' or entry.startswith('.'): continue
            proj_path = os.path.join(projects_dir, entry)
            if os.path.isdir(proj_path) and entry not in valid_project_names:
                tfvars_path = os.path.join(proj_path, 'terraform.tfvars')
                if os.path.exists(tfvars_path):
                    os.remove(tfvars_path)
                    print(f"⚠️  WARNING: Project '{entry}' was removed from SSoT (Excel).")
                    print(f"   -> Deleted 'terraform.tfvars' to exclude it from deployment.")
                    print(f"   -> Orphan directory 'terraform/4_projects/{entry}/' remains.")
                    print(f"   -> If GCP resources exist, destroy them first. Then run 'make prune' to delete the directory.")
                else:
                    print(f"ℹ️  INFO: Orphan directory found: 'terraform/4_projects/{entry}/' (not in SSoT, already excluded from deployment).")
                    print(f"   -> Run 'make prune' to remove it.")

    def is_true(val):
        if val is None: return False
        if isinstance(val, bool): return val
        return str(val).strip().lower() == 'true'

    folders_tf_path = os.path.join(os.path.dirname(__file__), '../3_folders/auto_folders.tf')
    with open(folders_tf_path, 'w') as f:
        f.write("# Auto-generated file. Do not edit manually.\n\n")
        for folder_name, parent_name in folders_map.items():
            fid = sanitize_id(folder_name)
            parent_str = str(parent_name).strip()
            parent_expr = "data.google_organization.org.name" if parent_str == 'organization_id' else f"google_folder.{sanitize_id(parent_str)}.name"
            f.write(f'resource "google_folder" "{fid}" {{\n  display_name = "{folder_name}"\n  parent = {parent_expr}\n  deletion_protection = false\n}}\n\n')
            f.write(f'output "{fid}_folder_id" {{\n  description = "The resource ID of the {folder_name} folder."\n  value = google_folder.{fid}.id\n}}\n\n')

    tags_tf_path = os.path.join(os.path.dirname(__file__), '../2_organization/auto_tags.tf')
    with open(tags_tf_path, 'w') as f:
        f.write("# Auto-generated file. Do not edit manually.\n\n")
        tag_value_map = {}
        for key, info in tag_definitions.items():
            kid = sanitize_id(key)
            f.write(
                f'resource "google_tags_tag_key" "{kid}" {{\n'
                f'  count = var.enable_tags ? 1 : 0\n'
                f'  parent = "organizations/${{data.google_organization.org.org_id}}"\n'
                f'  short_name = "{key}"\n'
                f'  description = "{info["description"]}"\n'
                f'}}\n\n'
            )
            for val in info['allowed_values']:
                vid = sanitize_id(f"{key}_{val}")
                f.write(
                    f'resource "google_tags_tag_value" "{vid}" {{\n'
                    f'  count = var.enable_tags ? 1 : 0\n'
                    f'  parent = try(google_tags_tag_key.{kid}[0].id, "")\n'
                    f'  short_name = "{val}"\n'
                    f'}}\n\n'
                )
                tag_value_map[f"{key}/{val}"] = f"try(google_tags_tag_value.{vid}[0].id, null)"
        f.write('output "tag_value_ids" {\n  description = "Map of organization tag key/value pairs to their resource IDs."\n  value = {\n')
        for k, v in tag_value_map.items(): f.write(f'    "{k}" = {v}\n')
        f.write('  }\n}\n\n')

    # --- ▼ 変更: マジックを排除したクリーンなCSV出力 ---
    def export_sheet_to_csv(sheet_name, output_path):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            header_row = [cell.value for cell in ws[1]]
            valid_col_indices = [i for i, h in enumerate(header_row) if h is not None and str(h).strip() != ""]
            if not valid_col_indices: return

            with open(output_path, 'w', newline='', encoding='utf-8') as csv_file:
                writer = csv.writer(csv_file)
                for row in ws.iter_rows(values_only=True):
                    if not any(cell is not None and str(cell).strip() != "" for cell in row): continue
                    filtered_row = [str(row[i]).strip() if i < len(row) and row[i] is not None else "" for i in valid_col_indices]
                    writer.writerow(filtered_row)

    export_sheet_to_csv('log_sinks', os.path.join(os.path.dirname(__file__), '../1_core/services/logsink/sinks/gcp_log_sink_config.csv'))
    export_sheet_to_csv('alert_definitions', os.path.join(os.path.dirname(__file__), '../1_core/services/monitoring/2_alert_policies/logsink_log_alerts/alert_definitions.csv'))
    export_sheet_to_csv('notifications', os.path.join(os.path.dirname(__file__), '../1_core/services/monitoring/1_notification_channels/notifications.csv'))
    export_sheet_to_csv('notifications', os.path.join(os.path.dirname(__file__), '../1_core/services/monitoring/2_alert_policies/logsink_log_alerts/notifications.csv'))

    subnets_tf_path = os.path.join(os.path.dirname(__file__), '../1_core/base/vpc-host/auto_subnets.tf')
    if 'shared_vpc_subnets' in wb.sheetnames:
        with open(subnets_tf_path, 'w') as f:
            f.write("# Auto-generated file. Do not edit manually.\n\n")
            ws = wb['shared_vpc_subnets']
            headers = [cell.value for cell in ws[1]]
            subnet_outputs = []
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row): continue
                s = dict(zip(headers, row))
                s_name = str(s.get('subnet_name') or '').strip()
                env = str(s.get('host_project_env') or '').strip().lower()
                cidr = str(s.get('ip_cidr_range') or '').strip()
                region = str(s.get('region') or '').strip()
                if s_name and env in ['prod', 'dev']:
                    sid = sanitize_id(s_name)
                    f.write(
                        f'resource "google_compute_subnetwork" "{sid}" {{\n'
                        f'  count = var.enable_vpc_host_projects ? 1 : 0\n'
                        f'  name = "{s_name}"\n'
                        f'  ip_cidr_range = "{cidr}"\n'
                        f'  region = "{region}"\n'
                        f'  network = try(google_compute_network.vpc_{env}[0].id, "")\n'
                        f'  project = try(module.vpc_host_{env}[0].project_id, "")\n'
                        f'  private_ip_google_access = true\n'
                        f'}}\n\n'
                    )
                    subnet_outputs.append(f'    "{s_name}" = try(google_compute_subnetwork.{sid}[0].id, null)')
            f.write('output "shared_vpc_subnet_ids" {\n  description = "Map of shared VPC subnet names to their resource IDs."\n  value = {\n')
            f.write('\n'.join(subnet_outputs) + '\n  }\n}\n\n')

    vpc_sc_tf_path = os.path.join(os.path.dirname(__file__), '../2_organization/auto_vpc_sc.tf')
    with open(vpc_sc_tf_path, 'w') as f:
        f.write("# Auto-generated file. Do not edit manually.\n\n")
        access_level_ids = {}
        if 'vpc_sc_access_levels' in wb.sheetnames:
            ws = wb['vpc_sc_access_levels']
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                al = dict(zip(headers, row))
                if not al.get('access_level_name'): continue
                sid = sanitize_id(al['access_level_name'])
                f.write(
                    f'resource "google_access_context_manager_access_level" "{sid}" {{\n'
                    f'  count = var.enable_vpc_sc ? 1 : 0\n'
                    f'  parent = "accessPolicies/${{try(google_access_context_manager_access_policy.access_policy[0].name, "")}}"\n'
                    f'  name = "accessPolicies/${{try(google_access_context_manager_access_policy.access_policy[0].name, "")}}/accessLevels/{al["access_level_name"]}"\n'
                    f'  title = "{al["access_level_name"]}"\n'
                    f'  basic {{\n'
                    f'    conditions {{\n'
                )
                if al.get('ip_subnetworks'):
                    ips = [ip.strip() + '/32' if '/' not in ip.strip() else ip.strip() for ip in str(al['ip_subnetworks']).split(',') if ip.strip()]
                    f.write(f'      ip_subnetworks = {json.dumps(ips)}\n')
                if al.get('members'):
                    members = [m.strip() for m in str(al['members']).split(',') if m.strip()]
                    f.write(f'      members = {json.dumps(members)}\n')
                f.write(f'    }}\n  }}\n}}\n\n')
                access_level_ids[al["access_level_name"]] = f"var.enable_vpc_sc ? google_access_context_manager_access_level.{sid}[0].name : null"
        perimeter_ids = {}
        for p in perimeters:
            if not p.get('perimeter_name'): continue
            sid = sanitize_id(p['perimeter_name'])
            services = [s.strip() for s in str(p.get('restricted_services') or '').split(',') if s.strip()]
            is_dry_run = is_true(p.get('dry_run'))
            f.write(
                f'resource "google_access_context_manager_service_perimeter" "{sid}" {{\n'
                f'  count = var.enable_vpc_sc ? 1 : 0\n'
                f'  parent = "accessPolicies/${{try(google_access_context_manager_access_policy.access_policy[0].name, "")}}"\n'
                f'  name = "accessPolicies/${{try(google_access_context_manager_access_policy.access_policy[0].name, "")}}/servicePerimeters/{p["perimeter_name"]}"\n'
                f'  title = "{p["perimeter_name"]}"\n'
            )
            if is_dry_run:
                f.write(f'  use_explicit_dry_run_spec = true\n')
                f.write(f'  spec {{\n')
                f.write(f'    restricted_services = {json.dumps(services)}\n')
                f.write(f'  }}\n')
                f.write(f'  lifecycle {{ ignore_changes = [status[0].resources, spec[0].resources] }}\n')
            else:
                f.write(f'  status {{\n')
                f.write(f'    restricted_services = {json.dumps(services)}\n')
                f.write(f'  }}\n')
                f.write(f'  lifecycle {{ ignore_changes = [status[0].resources] }}\n')
            f.write(f'}}\n\n')
            perimeter_ids[p['perimeter_name']] = f"var.enable_vpc_sc ? google_access_context_manager_service_perimeter.{sid}[0].name : null"
        f.write('output "service_perimeter_ids" {\n  description = "Map of VPC-SC perimeter names to their resource IDs."\n  value = {\n')
        for k, v in perimeter_ids.items(): f.write(f'    "{k}" = {v}\n')
        f.write('  }\n}\n\n')
        f.write('output "access_level_ids" {\n  description = "Map of VPC-SC access level names to their resource IDs."\n  value = {\n')
        for k, v in access_level_ids.items(): f.write(f'    "{k}" = {v}\n')
        f.write('  }\n}\n\n')

    if 'org_policies' in wb.sheetnames:
        org_policy_files = {}
        ws = wb['org_policies']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            p = dict(zip(headers, row))
            target_name = str(p.get('target_name') or '').strip()
            if not target_name: continue
            
            if target_name == 'organization_id':
                parent_resource = "organizations/${data.google_organization.org.org_id}"
                tf_file_path = "../2_organization/auto_org_policies.tf"
            elif target_name in folders_map:
                parent_resource = f"${{google_folder.{sanitize_id(target_name)}.name}}" # google_folder.name already includes "folders/" prefix
                tf_file_path = "../3_folders/auto_org_policies.tf"
            else:
                # Target is an application project
                parent_resource = f"projects/${{module.baseline.project_id}}"
                tf_file_path = f"../4_projects/{target_name}/auto_org_policies.tf"

            if tf_file_path not in org_policy_files:
                org_policy_files[tf_file_path] = []

            policy_id = str(p.get('policy_id') or '').strip()
            res_name = to_snake_case(f"{target_name}_{policy_id}")

            # ルール内容（ブール型 enforce / リスト型 allow_list）を組み立て、spec と dry_run_spec で共用する
            if is_true(p.get('enforce')):
                rules_block = '    rules {\n      enforce = "TRUE"\n    }\n'
            elif str(p.get('enforce')).strip().lower() == 'false':
                rules_block = '    rules {\n      enforce = "FALSE"\n    }\n'
            else:
                values = [v.strip() for v in str(p.get('allow_list') or '').split(',') if v.strip()]
                rules_block = f'    rules {{\n      values {{\n        allowed_values = {json.dumps(values)}\n      }}\n    }}\n'

            # apply_mode: live(本番=spec のみ) / dryrun(試行=dry_run_spec のみ) / both(両方を併記)。
            # 空欄・不正値は live にフォールバック（後方互換）。google_org_policy_policy は
            # 1リソース内に spec と dry_run_spec を同居できるため、リソースは1つにする（name 衝突を避ける）。
            apply_mode = str(p.get('apply_mode') or 'live').strip().lower()
            if apply_mode not in ('live', 'dryrun', 'both'):
                apply_mode = 'live'

            tf_block = f'resource "google_org_policy_policy" "{res_name}" {{\n'
            tf_block += f'  count  = var.enable_org_policies ? 1 : 0\n'
            tf_block += f'  name   = "{parent_resource}/policies/{policy_id}"\n'
            tf_block += f'  parent = "{parent_resource}"\n'
            if apply_mode in ('live', 'both'):
                tf_block += f'  spec {{\n{rules_block}  }}\n'
            if apply_mode in ('dryrun', 'both'):
                tf_block += f'  dry_run_spec {{\n{rules_block}  }}\n'
            tf_block += f'}}\n'
            org_policy_files[tf_file_path].append(tf_block)

        for path, blocks in org_policy_files.items():
            full_path = os.path.join(os.path.dirname(__file__), path)
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            with open(full_path, 'w') as f:
                f.write("# Auto-generated file. Do not edit manually.\n\n")
                f.write("\n".join(blocks))

    template_dir = os.path.join(os.path.dirname(__file__), '../4_projects/template')
    for proj in projects:
        app_name = str(proj.get('resource_name') or '').strip()
        if not app_name: continue
        project_dir = os.path.join(os.path.dirname(__file__), f"../4_projects/{app_name}")
        os.makedirs(project_dir, exist_ok=True)
        
        if os.path.exists(template_dir) and app_name != 'template':
            for filename in ['main.tf', 'variables.tf', 'provider.tf', 'versions.tf']:
                src = os.path.join(template_dir, filename)
                dst = os.path.join(project_dir, filename)
                if os.path.exists(src) and not os.path.exists(dst):
                    shutil.copy2(src, dst)
            backend_dst = os.path.join(project_dir, 'backend.tf')
            if not os.path.exists(backend_dst):
                with open(backend_dst, 'w') as f:
                    f.write(f'terraform {{\n  backend "gcs" {{\n    bucket = ""\n    prefix = "projects/{app_name}"\n  }}\n}}\n')

        parent_folder = str(proj.get('parent_name') or '').strip()
        # organization_id の場合は HCLの予約語 null を出力し、それ以外はダブルクォーテーションで囲む
        folder_id_val = "null" if parent_folder == 'organization_id' else f'"{sanitize_id(parent_folder)}"'
        # 既存プロジェクト採用(adopt)モード: existing_project_id があれば既存IDを採用（空なら新規作成）
        existing_pid = str(proj.get('existing_project_id') or '').strip()
        # environment は任意（空欄なら env ラベルを付けない）。検証は済み。
        env, _ = resolve_environment(proj.get('environment'))
        shared_vpc_env = resolve_shared_vpc_env(env, proj.get('shared_vpc'))

        # 課金アカウント(billing_account 列): 空欄→ 新規は module 側で global にリンク、
        # adopt(既存ID指定)は既存リンクを尊重して "manual"（TF は課金を管理しない）。
        # "manual" 明示 → TF 非管理。"<id>" → 指定アカウントにリンク。
        bill_col = str(proj.get('billing_account') or '').strip()
        billing_account_val = 'manual' if (bill_col == '' and existing_pid) else bill_col

        # --- ▼ 変更: ownerのマジック置換を排除（純粋な値を出力） ---
        tfvars_content = f"""# Auto-generated file. Do not edit manually.
organization_domain = "{domain}"
mgmt_project_id     = "{mgmt_project_id}"
app_name            = "{app_name}"
existing_project_id = "{existing_pid}"
billing_account     = "{billing_account_val}"
environment         = "{env}"
folder_id           = {folder_id_val}
shared_vpc_env      = "{shared_vpc_env}"
shared_vpc_subnet   = "{str(proj.get('shared_vpc') or '').strip()}"
vpc_sc              = "{str(proj.get('vpc_sc') or '').strip()}"
central_monitoring  = {str(is_true(proj.get('central_monitoring'))).lower()}
central_logging     = {str(is_true(proj.get('central_logging'))).lower()}
budget_amount       = {proj.get('budget_amount', 0) or 0}
budget_alert_emails = {json.dumps([e.strip() for e in str(proj.get('budget_alert_emails') or '').split(',') if e.strip()])}
org_tags            = {json.dumps([t.strip() for t in str(proj.get('org_tags') or '').split(',') if t.strip()])}
deletion_protection = true

labels = {{
  env   = "{env}"
  owner = "{str(proj.get('owner') or '').strip()}"
  app   = "{app_name}"
}}
"""
        with open(os.path.join(project_dir, 'terraform.tfvars'), 'w') as f: f.write(tfvars_content)

    # --- 監視ダッシュボード: central_monitoring=true のプロジェクトを 3_dashboards へ自動注入 ---
    # ダッシュボードのフィルタは実 GCP プロジェクトID（採用時は existing_project_id、
    # 新規は <prefix>-<app_name>）で限定するため、その実IDを列挙する。
    monitored_ids = []
    for proj in projects:
        if not is_true(proj.get('central_monitoring')):
            continue
        app_name = str(proj.get('resource_name') or '').strip()
        if not app_name:
            continue
        pid = str(proj.get('existing_project_id') or '').strip() or f"{project_id_prefix}-{app_name}"
        monitored_ids.append(pid)
    dashboards_dir = os.path.join(os.path.dirname(__file__), '../1_core/services/monitoring/3_dashboards')
    if os.path.isdir(dashboards_dir):
        with open(os.path.join(dashboards_dir, 'terraform.tfvars'), 'w') as f:
            f.write("# Auto-generated file. Do not edit manually.\n")
            f.write(f"monitored_project_ids = {json.dumps(monitored_ids)}\n")

    # --- Silence Undeclared Variable Warnings ---
    global_keys = [
        "terraform_service_account_email", "gcs_backend_bucket", "organization_domain",
        "gcp_region", "project_id_prefix", "enable_vpc_host_projects",
        "enable_shared_vpc", "enable_vpc_sc", "enable_org_policies", "enable_simplified_admin_groups",
        "enable_group_iam", "allow_resource_destruction", "enable_tags", "billing_account_id",
        "budget_threshold_percents", "focus_services"
    ]
    for root, dirs, files in os.walk(os.path.join(os.path.dirname(__file__), '../')):
        if any(f.endswith('.tf') for f in files) and '.terraform' not in root and 'modules' not in root:
            existing_vars = set()
            if 'variables.tf' in files:
                with open(os.path.join(root, 'variables.tf'), 'r') as vf:
                    existing_vars.update(re.findall(r'variable\s+"([^"]+)"', vf.read()))
            
            missing_vars = [k for k in global_keys if k not in existing_vars]
            if missing_vars:
                with open(os.path.join(root, 'auto_global_vars.tf'), 'w') as fv:
                    fv.write('# Auto-generated file to silence undeclared variable warnings.\n')
                    for mv in missing_vars:
                        fv.write(f'variable "{mv}" {{\n  type        = any\n  description = "Auto-generated dummy variable to silence tflint warnings."\n  default     = null\n}}\n\n')

    print(f"✅ Generated all resources successfully")

if __name__ == "__main__":
    generate_resources()
