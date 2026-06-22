# /// script
# requires-python = ">=3.12"
# dependencies = [
#     "openpyxl==3.1.2",
# ]
# ///
"""
納品物（構築設定明細書）生成スクリプト。

SSoT である gcp-foundations.xlsx と common.tfvars / domain.env から、
日本のシステム開発における一般的な「設計・設定明細書」様式の Excel を生成する。
`make delivery`（handover）の直前に実行され、生成物は delivery/ 配下に出力される。

メタ情報（顧客名・作成者・版数等）は環境変数で上書き可能:
  DELIVERY_CUSTOMER / DELIVERY_VENDOR / DELIVERY_AUTHOR / DELIVERY_VERSION / DELIVERY_DOCNO
"""
import os
import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------------------
# パス設定
# ------------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, "../.."))
XLSX_PATH = os.path.join(ROOT_DIR, "gcp-foundations.xlsx")
COMMON_TFVARS = os.path.join(ROOT_DIR, "terraform", "common.tfvars")
DOMAIN_ENV = os.path.join(ROOT_DIR, "domain.env")
OUTPUT_DIR = os.path.join(ROOT_DIR, "delivery")

# ------------------------------------------------------------------------------
# スタイル定義（日本のベンダー様式に寄せた配色・罫線）
# ------------------------------------------------------------------------------
FONT_NAME = "Meiryo"
C_HEADER_BG = "1F3864"   # 濃紺（表ヘッダ）
C_HEADER_FG = "FFFFFF"
C_SECTION_BG = "2E5395"  # セクション見出し
C_LABEL_BG = "D9E1F2"    # ラベル列（薄い青）
C_TITLE_FG = "1F3864"
C_DONE = "375623"        # 「完了」文字色（緑）

THIN = Side(style="thin", color="9CA3AF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

F_TITLE = Font(name=FONT_NAME, size=22, bold=True, color=C_TITLE_FG)
F_SUBTITLE = Font(name=FONT_NAME, size=12, color="595959")
F_SECTION = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
F_HEADER = Font(name=FONT_NAME, size=10, bold=True, color=C_HEADER_FG)
F_LABEL = Font(name=FONT_NAME, size=10, bold=True, color="1F3864")
F_BODY = Font(name=FONT_NAME, size=10, color="000000")
F_DONE = Font(name=FONT_NAME, size=10, bold=True, color=C_DONE)
F_NOTE = Font(name=FONT_NAME, size=9, color="808080", italic=True)

FILL_HEADER = PatternFill("solid", fgColor=C_HEADER_BG)
FILL_SECTION = PatternFill("solid", fgColor=C_SECTION_BG)
FILL_LABEL = PatternFill("solid", fgColor=C_LABEL_BG)

AL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
AL_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ------------------------------------------------------------------------------
# 入力読み込み
# ------------------------------------------------------------------------------
def read_tfvars(path):
    """common.tfvars からトップレベルのスカラー値を素朴に読む（key = value）。"""
    data = {}
    if not os.path.exists(path):
        return data
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key = key.strip()
            val = val.split("#", 1)[0].strip().strip('"').strip()
            # リスト/マップ行はスキップ（複数行構造は対象外）
            if val.startswith(("[", "{")) or not key.replace("_", "").isalnum():
                continue
            data[key] = val
    return data


def read_domain(path):
    if not os.path.exists(path):
        return ""
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            if line.strip().startswith("domain="):
                return line.split("=", 1)[1].strip().strip('"')
    return ""


def read_sheet(wb, name):
    """シートを (headers, [dict,...]) で返す。無ければ ([], [])。"""
    if name not in wb.sheetnames:
        return [], []
    ws = wb[name]
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return headers, rows


# ------------------------------------------------------------------------------
# シート構築ヘルパ
# ------------------------------------------------------------------------------
def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def section_title(ws, row, text, span):
    """セクション見出し（塗りつぶし帯）を row に書き、次の行番号を返す。"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SECTION
    c.fill = FILL_SECTION
    c.alignment = AL_LEFT
    ws.row_dimensions[row].height = 22
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = FILL_SECTION
    return row + 1


def table(ws, row, headers, records, col_keys=None, done_col=None):
    """表ヘッダ＋明細を描画し、次の行番号を返す。
    headers: 表示用ヘッダ名リスト
    records: dict のリスト
    col_keys: 各列に対応する dict キー（None なら headers と同じ）
    done_col: 「完了」など緑表示にする列キー
    """
    col_keys = col_keys or headers
    span = len(headers)
    # ヘッダ行
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.alignment = AL_CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 20
    row += 1
    # 明細行
    if not records:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        c = ws.cell(row=row, column=1, value="（定義なし）")
        c.font = F_NOTE
        c.alignment = AL_CENTER
        for col in range(1, span + 1):
            ws.cell(row=row, column=col).border = BORDER
        return row + 2
    for rec in records:
        for j, key in enumerate(col_keys, start=1):
            val = rec.get(key, "")
            if val is None:
                val = ""
            c = ws.cell(row=row, column=j, value=val)
            c.border = BORDER
            c.alignment = AL_TOP
            c.font = F_DONE if (done_col and key == done_col and str(val) == "完了") else F_BODY
        row += 1
    return row + 1


def kv_table(ws, row, pairs, label_w=2):
    """ラベル/値の2列（または span 結合値）テーブル。pairs=[(label, value), ...]。"""
    for label, value in pairs:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = F_LABEL
        lc.fill = FILL_LABEL
        lc.alignment = AL_LEFT
        lc.border = BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=label_w + 2)
        vc = ws.cell(row=row, column=2, value="" if value is None else value)
        vc.font = F_BODY
        vc.alignment = AL_LEFT
        for col in range(1, label_w + 3):
            ws.cell(row=row, column=col).border = BORDER
        row += 1
    return row + 1


def yn(flag):
    return "有効" if str(flag).lower() == "true" else "無効"


# ------------------------------------------------------------------------------
# 各シート生成
# ------------------------------------------------------------------------------
def build_cover(wb, meta):
    ws = wb.create_sheet("表紙")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [4, 22, 30, 16, 16])

    ws.cell(row=3, column=2, value="Google Cloud Platform").font = F_SUBTITLE
    t = ws.cell(row=5, column=2, value="基盤構築 設定明細書")
    t.font = F_TITLE
    ws.merge_cells("B5:E6")
    # 22pt のタイトルが上側で見切れないよう、結合した5・6行目の高さを確保する
    ws.row_dimensions[5].height = 19
    ws.row_dimensions[6].height = 19
    ws.cell(row=8, column=2, value="GCP Foundations 納品ドキュメント").font = F_SUBTITLE

    pairs = [
        ("顧客名", meta["customer"]),
        ("対象組織ドメイン", meta["org_domain"]),
        ("組織 ID", meta["org_id"]),
        ("文書番号", meta["doc_no"]),
        ("版数", meta["version"]),
        ("作成日", meta["date"]),
        ("作成者", meta["author"]),
        ("提供元", meta["vendor"]),
    ]
    row = 12
    for label, value in pairs:
        lc = ws.cell(row=row, column=2, value=label)
        lc.font = F_LABEL
        lc.fill = FILL_LABEL
        lc.alignment = AL_LEFT
        lc.border = BORDER
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        vc = ws.cell(row=row, column=3, value=value)
        vc.font = F_BODY
        vc.alignment = AL_LEFT
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = BORDER
        ws.row_dimensions[row].height = 22
        row += 1

    # 承認欄（押印枠）
    row += 1
    ws.cell(row=row, column=2, value="承認").font = F_LABEL
    row += 1
    appr_headers = ["承認", "確認", "作成"]
    for j, h in enumerate(appr_headers, start=3):
        c = ws.cell(row=row, column=j, value=h)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.alignment = AL_CENTER
        c.border = BORDER
    lc = ws.cell(row=row, column=2, value="役職／氏名")
    lc.font = F_LABEL
    lc.fill = FILL_LABEL
    lc.border = BORDER
    lc.alignment = AL_CENTER
    for r in range(row + 1, row + 3):
        ws.cell(row=r, column=2).border = BORDER
        ws.row_dimensions[r].height = 30
        for j in range(3, 6):
            ws.cell(row=r, column=j).border = BORDER


def build_revision(wb, meta):
    ws = wb.create_sheet("改訂履歴")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [10, 16, 50, 18, 14])
    row = section_title(ws, 2, "改訂履歴", 5)
    records = [{
        "版数": meta["version"],
        "改訂日": meta["date"],
        "改訂内容": "初版作成（GCP 基盤構築の納品）",
        "作成者": meta["author"],
        "承認": "",
    }]
    table(ws, row, ["版数", "改訂日", "改訂内容", "作成者", "承認"], records)


def build_toc(wb):
    ws = wb.create_sheet("目次")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [8, 52])
    row = section_title(ws, 2, "目次", 2)
    entries = [
        ("1", "構築項目一覧"),
        ("2", "構築概要（環境情報）"),
        ("3", "フォルダ構成"),
        ("4", "プロジェクト一覧"),
        ("5", "組織ポリシー設定一覧"),
        ("6", "ログ集約シンク設定"),
        ("7", "BigQuery リソース"),
        ("8", "監視・予算（集中モニタリング／アラート／通知／予算）"),
        ("9", "ネットワーク（Shared VPC／VPC-SC）"),
        ("10", "タグ・ラベル"),
        ("11", "Google グループ・IAM"),
        ("12", "費用に関する注意事項"),
    ]
    records = [{"章": no, "項目": title} for no, title in entries]
    table(ws, row, ["章", "項目"], records, col_keys=["章", "項目"])


def build_summary(wb, meta, ctx):
    """構築項目一覧（実施サマリ）— 顧客が一目で「何を実施したか」を確認できる総括表。"""
    ws = wb.create_sheet("1.構築項目一覧")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [5, 30, 46, 12])
    row = section_title(ws, 2, "1. 構築項目一覧", 4)

    items = [
        ("組織への取り込み", "組織なしプロジェクトを GCP 組織配下へ移管し Terraform 管理下に統制",
         "完了" if ctx["adopted_projects"] else "—"),
        ("Cloud Identity 認証", "Cloud Identity による組織アカウント／ID 認証基盤の整備", "完了"),
        ("請求先アカウントのリンク", "対象プロジェクトと請求先アカウントの紐付け", "完了"),
        ("管理フォルダ構成", "admin／network 等の管理用フォルダ階層の構築",
         "完了" if ctx["structural_folders"] else "—"),
        ("管理プロジェクト構築", "ログ集約・モニタリング用の管理プロジェクト作成",
         "完了" if ctx["mgmt_projects"] else "—"),
        ("組織ポリシー設定", "セキュア・バイ・デフォルトの組織ポリシー適用",
         "完了" if ctx["org_policies"] else "—"),
        ("ログ集約シンク設定", "監査ログ等を集約ログプロジェクト(BigQuery等)へエクスポート",
         "完了" if ctx["log_sinks"] else "—"),
        ("集中モニタリング設定", "対象プロジェクトをモニタリング対象スコープへ登録",
         "完了" if ctx["monitoring_targets"] else "—"),
        ("予算アラート設定", "予算しきい値超過時のアラート通知設定",
         "完了" if ctx["budget_projects"] else "—"),
        ("ネットワーク構成", "Shared VPC / VPC Service Controls の構成",
         "完了" if (ctx["subnets"] or ctx["perimeters"]) else "—"),
        ("タグ・ラベル付与", "組織タグの定義と、プロジェクトへのラベル(app/env/owner)付与",
         "完了" if (ctx["projects"] or ctx["tag_definitions"]) else "—"),
        ("Google グループ・IAM", "管理用 Google グループへの組織レベル IAM ロール付与",
         "完了" if str(ctx["tfvars"].get("enable_group_iam", "true")).lower() == "true" else "—"),
    ]
    records = [{"No": i, "構築項目": n, "概要": d, "実施状況": s}
               for i, (n, d, s) in enumerate(items, start=1)]
    row = table(ws, row, ["No", "構築項目", "概要", "実施状況"], records,
                col_keys=["No", "構築項目", "概要", "実施状況"], done_col="実施状況")
    c = ws.cell(row=row, column=1,
                value="※「—」は本案件の SSoT（gcp-foundations.xlsx）に定義がない項目です。")
    c.font = F_NOTE


def build_overview(wb, meta, ctx):
    ws = wb.create_sheet("2.構築概要")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [22, 30, 18, 18])
    row = section_title(ws, 2, "2. 構築概要（環境情報）", 4)
    tv = ctx["tfvars"]
    pairs = [
        ("組織ドメイン", meta["org_domain"]),
        ("組織 ID", meta["org_id"]),
        ("プロジェクト ID 接頭辞", tv.get("project_id_prefix", "")),
        ("請求先アカウント", tv.get("billing_account_id", "")),
        ("既定リージョン", tv.get("default_region", tv.get("gcp_region", tv.get("region", "")))),
        ("Shared VPC", yn(tv.get("enable_shared_vpc", "false"))),
        ("VPC ホストプロジェクト", yn(tv.get("enable_vpc_host_projects", "false"))),
        ("組織ポリシー適用", yn(tv.get("enable_org_policies", "true"))),
        ("組織タグ", yn(tv.get("enable_tags", "false"))),
        ("リソース削除許可", yn(tv.get("allow_resource_destruction", "false"))),
    ]
    row = kv_table(ws, row, pairs, label_w=2)

    row = section_title(ws, row, "管理プロジェクト一覧（基盤標準）", 4)
    records = [{"用途": p[0], "プロジェクト ID（規約）": p[1], "配置先フォルダ": p[2], "説明": p[3]}
               for p in ctx["mgmt_projects"]]
    table(ws, row, ["用途", "プロジェクト ID（規約）", "配置先フォルダ", "説明"], records,
          col_keys=["用途", "プロジェクト ID（規約）", "配置先フォルダ", "説明"])


def build_folders(wb, ctx):
    ws = wb.create_sheet("3.フォルダ構成")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [16, 24, 22, 40])
    row = section_title(ws, 2, "3. フォルダ構成", 4)
    records = []
    for f in ctx["structural_folders"]:
        records.append({"区分": "基盤標準(L0)", "フォルダ名": f["resource_name"],
                        "親": f["parent_name"], "用途": f["purpose"]})
    for f in ctx["folders"]:
        records.append({"区分": "SSoT定義", "フォルダ名": f["resource_name"],
                        "親": f["parent_name"], "用途": ""})
    row = table(ws, row, ["区分", "フォルダ名", "親（フォルダ／組織）", "用途"], records,
                col_keys=["区分", "フォルダ名", "親", "用途"])
    c = ws.cell(row=row, column=1,
                value="※ admin／network は基盤標準フォルダ（terraform/0_bootstrap で組織直下に常時作成）。"
                      "admin に管理プロジェクト（logsink／monitoring）、network に Shared VPC ホストを配置。"
                      "それ以外は SSoT(gcp-foundations.xlsx) の resources シート(resource_type=folder)で定義。")
    c.font = F_NOTE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)


def build_projects(wb, ctx):
    ws = wb.create_sheet("4.プロジェクト一覧")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [22, 18, 10, 26, 12, 12, 14, 14])

    # 4-1. 管理プロジェクト（基盤標準・SSoT 外。1_core/base で作成）
    row = section_title(ws, 2, "4-1. 管理プロジェクト（基盤標準）", 8)
    mgmt_records = [{"用途": m[0], "プロジェクト ID（規約）": m[1],
                     "配置先フォルダ": m[2], "説明": m[3]} for m in ctx["mgmt_projects"]]
    row = table(ws, row, ["用途", "プロジェクト ID（規約）", "配置先フォルダ", "説明"], mgmt_records,
                col_keys=["用途", "プロジェクト ID（規約）", "配置先フォルダ", "説明"])

    # 4-2. アプリケーションプロジェクト（SSoT の resources シート由来）
    row = section_title(ws, row, "4-2. アプリケーションプロジェクト（SSoT）", 8)
    records = []
    for p in ctx["projects"]:
        records.append({
            "表示名": str(p.get("resource_name") or ""),
            "配置先": str(p.get("parent_name") or ""),
            "環境": str(p.get("environment") or ""),
            "既存ID(取り込み)": str(p.get("existing_project_id") or ""),
            "集中監視": "○" if str(p.get("central_monitoring")).lower() == "true" else "",
            "集中ログ": "○" if str(p.get("central_logging")).lower() == "true" else "",
            "予算(円)": p.get("budget_amount") or "",
            "予算通知先": str(p.get("budget_alert_emails") or ""),
        })
    table(ws, row,
          ["表示名", "配置先", "環境", "既存ID(取り込み)", "集中監視", "集中ログ", "予算(円)", "予算通知先"],
          records,
          col_keys=["表示名", "配置先", "環境", "既存ID(取り込み)", "集中監視", "集中ログ", "予算(円)", "予算通知先"])


def build_org_policies(wb, ctx):
    ws = wb.create_sheet("5.組織ポリシー")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [22, 40, 12, 30, 12])
    row = section_title(ws, 2, "5. 組織ポリシー設定一覧", 5)
    records = []
    for p in ctx["org_policies"]:
        records.append({
            "適用対象": str(p.get("target_name") or ""),
            "ポリシーID": str(p.get("policy_id") or ""),
            "強制": "ON" if str(p.get("enforce")).lower() == "true" else "OFF",
            "許可リスト": str(p.get("allow_list") or ""),
            "適用モード": str(p.get("apply_mode") or "live"),
        })
    table(ws, row, ["適用対象", "ポリシーID", "強制", "許可リスト", "適用モード"], records,
          col_keys=["適用対象", "ポリシーID", "強制", "許可リスト", "適用モード"])


def build_log_sinks(wb, ctx):
    ws = wb.create_sheet("6.ログ集約シンク")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [20, 40, 16, 30, 12])
    row = section_title(ws, 2, "6. ログ集約シンク設定", 5)
    records = []
    for s in ctx["log_sinks"]:
        records.append({
            "ログ種別": str(s.get("log_type") or ""),
            "フィルタ": str(s.get("filter") or ""),
            "宛先種別": str(s.get("destination_type") or ""),
            "宛先": str(s.get("destination_parent") or ""),
            "保持日数": s.get("retention_days") or "",
        })
    table(ws, row, ["ログ種別", "フィルタ", "宛先種別", "宛先", "保持日数"], records,
          col_keys=["ログ種別", "フィルタ", "宛先種別", "宛先", "保持日数"])


def build_bigquery(wb, ctx):
    ws = wb.create_sheet("7.BigQuery リソース")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [28, 16, 16, 12, 40])
    tv = ctx["tfvars"]
    prefix = tv.get("project_id_prefix", "") or "<prefix>"
    logsink_project = f"{prefix}-logsink"
    region = tv.get("default_region", tv.get("gcp_region", tv.get("region", "")))

    # 7-1. ログシンクの BigQuery 宛先データセット（SSoT の log_sinks 由来）
    row = section_title(ws, 2, "7-1. ログシンク宛先データセット", 5)
    records = []
    for s in ctx["log_sinks"]:
        if str(s.get("destination_type") or "").strip().lower() != "bigquery":
            continue
        records.append({
            "データセット": str(s.get("destination_parent") or ""),
            "プロジェクト": logsink_project,
            "ロケーション": region,
            "保持日数": s.get("retention_days") or "",
            "用途（ログ種別）": str(s.get("log_type") or ""),
        })
    row = table(ws, row, ["データセット", "プロジェクト", "ロケーション", "保持日数", "用途（ログ種別）"], records,
                col_keys=["データセット", "プロジェクト", "ロケーション", "保持日数", "用途（ログ種別）"])
    c = ws.cell(row=row, column=1,
                value="※ log_sinks の BigQuery 宛先。パーティション表として作成され、保持日数はテーブル既定の有効期限に対応。"
                      "シンクの writer identity に roles/bigquery.dataEditor を付与済み。")
    c.font = F_NOTE
    row += 2

    # 7-2. アセットインベントリ（IAM 監査）— 基盤標準で構築される BigQuery リソース
    row = section_title(ws, row, "7-2. アセットインベントリ（IAM 監査・基盤標準）", 5)
    ai_records = [
        {"名前": "asset_inventory", "種別": "データセット", "プロジェクト": logsink_project, "形式": "—",
         "内容": "Cloud Asset Inventory(IAM_POLICY) リアルタイムフィードの宛先データセット"},
        {"名前": "asset_inventory.iam_policy", "種別": "テーブル", "プロジェクト": logsink_project, "形式": "JSON",
         "内容": "組織／フォルダ／プロジェクトの IAM ポリシー変更を追記（append-only 履歴）"},
        {"名前": "asset_inventory.v_iam_policy", "種別": "ビュー", "プロジェクト": logsink_project, "形式": "SQL",
         "内容": "IAM を resource × role × members に展開した監査用ビュー"},
    ]
    row = table(ws, row, ["名前", "種別", "プロジェクト", "形式", "内容"], ai_records,
                col_keys=["名前", "種別", "プロジェクト", "形式", "内容"])
    c = ws.cell(row=row, column=1,
                value="※ 基盤標準構成（asset_inventory_bq_export レイヤー）。CAI→Pub/Sub→BigQuery の準リアルタイム取込で、"
                      "Terraform 管理外の手動付与を含む IAM 付与状況を監査証跡として保持する。")
    c.font = F_NOTE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)


def build_monitoring(wb, ctx):
    ws = wb.create_sheet("8.監視・予算")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [24, 30, 36])
    tv = ctx["tfvars"]
    prefix = tv.get("project_id_prefix", "") or "<prefix>"
    monitoring_project = f"{prefix}-monitoring"

    # 8-1. 集中モニタリング設定（スコーピングプロジェクト・メトリクススコープ・ダッシュボード）
    row = section_title(ws, 2, "8-1. 集中モニタリング設定", 3)
    pairs = [
        ("スコーピングプロジェクト", f"{monitoring_project}（admin フォルダ配下）"),
        ("メトリクススコープ", f"locations/global/metricsScopes/{monitoring_project}"),
        ("監視対象の登録方式", "各プロジェクトの central_monitoring=true により自動登録"
                          "（google_monitoring_monitored_project）"),
        ("API ヘルスダッシュボード", "監視対象プロジェクトに対し作成（3_dashboards レイヤー／"
                              "Consumed API・Gemini API 中心・全14タイル）"),
    ]
    row = kv_table(ws, row, pairs, label_w=2)

    row = section_title(ws, row, "監視対象プロジェクト（メトリクススコープ登録）", 3)
    records = [{
        "プロジェクト": str(p.get("resource_name") or ""),
        "既存ID(取り込み)": str(p.get("existing_project_id") or ""),
        "ダッシュボード": "API ヘルス（3_dashboards）",
    } for p in ctx["monitoring_targets"]]
    row = table(ws, row, ["プロジェクト", "既存ID(取り込み)", "API ヘルスダッシュボード"], records,
                col_keys=["プロジェクト", "既存ID(取り込み)", "ダッシュボード"])
    c = ws.cell(row=row, column=1,
                value="※ central_monitoring=true のプロジェクトを集中モニタリングのメトリクススコープへ登録。"
                      "API ヘルスダッシュボードは 3_dashboards レイヤーの monitored_project_ids に指定した"
                      "プロジェクトに作成され、Gemini API 等の Consumed API を可視化する。")
    c.font = F_NOTE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    # 8-2. アラート定義
    row = section_title(ws, row, "8-2. アラート定義", 3)
    records = [{
        "アラート名": str(a.get("alert_display_name") or a.get("alert_name") or ""),
        "条件(metric filter)": str(a.get("metric_filter") or ""),
        "説明": str(a.get("alert_documentation") or ""),
    } for a in ctx["alert_defs"]]
    row = table(ws, row, ["アラート名", "条件(metric filter)", "説明"], records,
                col_keys=["アラート名", "条件(metric filter)", "説明"])

    row = section_title(ws, row, "8-3. 通知先", 3)
    records = [{
        "アラート名": str(n.get("alert_name") or ""),
        "通知先メール": str(n.get("user_email") or ""),
        "受信": "受信する" if str(n.get("receive_alerts")).lower() == "true" else "受信しない",
    } for n in ctx["notifications"]]
    row = table(ws, row, ["アラート名", "通知先メール", "受信可否"], records,
                col_keys=["アラート名", "通知先メール", "受信"])

    row = section_title(ws, row, "8-4. 予算アラート", 3)
    records = []
    for p in ctx["budget_projects"]:
        records.append({
            "プロジェクト": str(p.get("resource_name") or ""),
            "予算額(円)": p.get("budget_amount") or "",
            "通知先": str(p.get("budget_alert_emails") or ""),
        })
    table(ws, row, ["プロジェクト", "予算額(円)", "通知先メール"], records,
          col_keys=["プロジェクト", "予算額(円)", "通知先"])


def build_network(wb, ctx):
    ws = wb.create_sheet("9.ネットワーク")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [20, 24, 20, 24])

    row = section_title(ws, 2, "9-1. Shared VPC サブネット", 4)
    records = [{
        "ホスト環境": str(s.get("host_project_env") or ""),
        "サブネット名": str(s.get("subnet_name") or ""),
        "リージョン": str(s.get("region") or ""),
        "CIDR": str(s.get("ip_cidr_range") or ""),
    } for s in ctx["subnets"]]
    row = table(ws, row, ["ホスト環境", "サブネット名", "リージョン", "CIDR"], records,
                col_keys=["ホスト環境", "サブネット名", "リージョン", "CIDR"])

    row = section_title(ws, row, "9-2. VPC Service Controls 境界", 4)
    records = [{
        "境界名": str(p.get("perimeter_name") or ""),
        "タイトル": str(p.get("title") or ""),
        "制限サービス": str(p.get("restricted_services") or ""),
        "モード": "ドライラン" if str(p.get("dry_run")).lower() == "true" else "適用",
    } for p in ctx["perimeters"]]
    table(ws, row, ["境界名", "タイトル", "制限サービス", "ドライラン/適用"], records,
          col_keys=["境界名", "タイトル", "制限サービス", "モード"])


def build_tags_labels(wb, ctx):
    ws = wb.create_sheet("10.タグ・ラベル")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [24, 30, 40])

    row = section_title(ws, 2, "10-1. 組織タグ定義一覧", 3)
    records = [{
        "タグキー": str(t.get("tag_key") or ""),
        "許可値": str(t.get("allowed_values") or ""),
        "説明": str(t.get("description") or ""),
    } for t in ctx["tag_definitions"]]
    row = table(ws, row, ["タグキー", "許可値", "説明"], records,
                col_keys=["タグキー", "許可値", "説明"])
    c = ws.cell(row=row, column=1,
                value="※ 組織レベルのタグ。common.tfvars の enable_tags 有効時に組織へ作成され、"
                      "各プロジェクトの org_tags 指定に基づいて紐付けられます。")
    c.font = F_NOTE
    row += 2

    row = section_title(ws, row, "10-2. プロジェクト別 ラベル／適用タグ", 3)
    records = []
    for p in ctx["projects"]:
        app = str(p.get("resource_name") or "")
        env = str(p.get("environment") or "").strip().lower()
        owner = str(p.get("owner") or "").strip()
        tags = str(p.get("org_tags") or "").strip()
        labels = f"app={app}"
        if env:
            labels += f", env={env}"
        if owner:
            labels += f", owner={owner}"
        records.append({
            "プロジェクト": app,
            "ラベル (labels)": labels,
            "組織タグ (org_tags)": tags or "（なし）",
        })
    row = table(ws, row, ["プロジェクト", "ラベル (labels)", "組織タグ (org_tags)"], records,
                col_keys=["プロジェクト", "ラベル (labels)", "組織タグ (org_tags)"])
    c = ws.cell(row=row, column=1,
                value="※ ラベルは GCP プロジェクトの labels。app は必ず付与され、env／owner は SSoT で"
                      "値がある場合のみ付与されます（空欄は付与なし）。")
    c.font = F_NOTE


# 管理用 Google グループと組織レベル IAM ロール（terraform/2_organization/main.tf の
# locals.raw_roles を転記）。SSoT(Excel) には存在しないため、ここで定義を反映する。
GROUP_ROLES = {
    "gcp-organization-admins": (
        "組織管理者", [
            "roles/cloudkms.admin", "roles/cloudsupport.admin", "roles/iam.organizationRoleAdmin",
            "roles/orgpolicy.policyAdmin", "roles/pubsub.admin", "roles/resourcemanager.folderAdmin",
            "roles/resourcemanager.organizationAdmin", "roles/resourcemanager.projectCreator",
            "roles/securitycenter.admin",
        ]),
    "gcp-billing-admins": (
        "請求管理者", [
            "roles/billing.creator", "roles/resourcemanager.organizationViewer",
        ]),
    "gcp-vpc-network-admins": (
        "VPC ネットワーク管理者", [
            "roles/compute.networkAdmin", "roles/compute.securityAdmin", "roles/compute.xpnAdmin",
            "roles/resourcemanager.folderViewer",
        ]),
    "gcp-hybrid-connectivity-admins": (
        "ハイブリッド接続管理者", [
            "roles/compute.networkAdmin", "roles/resourcemanager.folderViewer",
        ]),
    "gcp-logging-monitoring-admins": (
        "ログ・監視 管理者", [
            "roles/logging.admin", "roles/monitoring.admin", "roles/pubsub.admin",
        ]),
    "gcp-logging-monitoring-viewers": (
        "ログ・監視 閲覧者", [
            "roles/logging.viewer", "roles/monitoring.viewer",
        ]),
    "gcp-security-admins": (
        "セキュリティ管理者", [
            "roles/cloudkms.admin", "roles/compute.viewer", "roles/container.viewer",
            "roles/iam.organizationRoleViewer", "roles/iam.securityAdmin", "roles/iam.securityReviewer",
            "roles/iam.serviceAccountCreator", "roles/logging.admin", "roles/logging.configWriter",
            "roles/logging.privateLogViewer", "roles/monitoring.admin", "roles/orgpolicy.policyAdmin",
            "roles/resourcemanager.folderIamAdmin", "roles/securitycenter.admin",
        ]),
    "gcp-developers": (
        "開発者", [
            "roles/browser", "roles/viewer", "roles/resourcemanager.organizationViewer",
        ]),
    "gcp-devops": (
        "DevOps", [
            "roles/resourcemanager.folderViewer",
        ]),
}


def resolve_group_roles(simplified):
    """terraform の group_roles と同じ選択ロジック。
    集約モード(simplified=True): 請求以外の全ロールを組織管理者へ統合し、2グループに集約。
    フルモード: 9グループをそのまま使用。返り値 [(group, desc, [roles]), ...]。
    """
    if not simplified:
        return [(g, d, r) for g, (d, r) in GROUP_ROLES.items()]
    merged, seen = [], set()
    for g, (d, roles) in GROUP_ROLES.items():
        if g == "gcp-billing-admins":
            continue
        for r in roles:
            if r not in seen:
                seen.add(r)
                merged.append(r)
    billing = GROUP_ROLES["gcp-billing-admins"]
    return [
        ("gcp-organization-admins", "組織管理者（集約モード：請求以外の全ロールを統合）", merged),
        ("gcp-billing-admins", billing[0], list(billing[1])),
    ]


def build_groups_iam(wb, ctx):
    ws = wb.create_sheet("11.Googleグループ・IAM")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [34, 26, 12, 52])
    tv = ctx["tfvars"]
    domain = ctx["org_domain"]
    simplified = str(tv.get("enable_simplified_admin_groups", "false")).lower() == "true"
    group_iam_on = str(tv.get("enable_group_iam", "true")).lower() == "true"

    row = section_title(ws, 2, "11. 管理用 Google グループと IAM ロール", 4)
    pairs = [
        ("グループ構成モード", "集約モード（2グループ）" if simplified else "フルモード（9グループ）"),
        ("組織レベル IAM 付与", "有効（適用済み）" if group_iam_on else "無効（IAM バインディング未適用）"),
        ("グループ メール形式", f"<グループ名>@{domain}"),
        ("付与レベル", "組織レベル（Organization IAM）"),
    ]
    row = kv_table(ws, row, pairs, label_w=2)

    row = section_title(ws, row, "グループ別 付与ロール一覧", 4)
    records = []
    for g, desc, roles in resolve_group_roles(simplified):
        records.append({
            "グループ（メール）": f"{g}@{domain}",
            "役割": desc,
            "付与レベル": "組織",
            "IAM ロール": "\n".join(roles),
        })
    row = table(ws, row, ["グループ（メール）", "役割", "付与レベル", "IAM ロール"], records,
                col_keys=["グループ（メール）", "役割", "付与レベル", "IAM ロール"])

    note = ("※ Google グループ自体は Cloud Identity / 管理コンソール（Cloud セットアップ）で事前作成します"
            "（Terraform はグループを作成せず、組織レベルの IAM ロール付与のみを管理）。")
    if not group_iam_on:
        note += " 現在は enable_group_iam=false のため、上記ロールの自動付与は未適用です（手動付与または将来適用）。"
    c = ws.cell(row=row, column=1, value=note)
    c.font = F_NOTE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)


def build_cost_notes(wb, ctx):
    ws = wb.create_sheet("12.費用注意事項")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [22, 40, 26, 40])
    row = section_title(ws, 2, "12. 費用に関する注意事項", 4)

    intro = ws.cell(row=row, column=1,
                    value="本基盤の構成は、設定内容や対象範囲によってGoogle Cloudの利用料金が増加する場合があります。"
                          "特にログ・監視まわりはデータ量に比例して課金されるため、以下を運用開始前にご確認ください。"
                          "（金額は利用量に依存します。詳細は Google Cloud の料金表・料金計算ツールをご参照ください。）")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    intro.font = F_BODY
    intro.alignment = AL_TOP
    ws.row_dimensions[row].height = 46
    row += 2

    notes = [
        ("データアクセス監査ログの集約",
         "Data Access 監査ログは出力量が非常に多く、ログ集約シンク経由で BigQuery / Cloud Storage に取り込むと保管・取り込み課金が大きく増加します。",
         "BigQuery 取り込み・保管 / Cloud Storage 保管",
         "既定では管理アクティビティ監査ログのみ集約。Data Access は必要なサービス・プロジェクトに限定して有効化する。"),
        ("VPC フローログ",
         "サンプリング率・集約間隔の設定によりログ生成量が大きく変動し、生成およびエクスポート（シンク先）の課金が増加します。",
         "フローログ生成 / シンク先の取り込み・保管",
         "必要なサブネットのみ有効化。サンプリング率を下げ、集約間隔を広げる。保持期間を短く設定する。"),
        ("ログ集約シンク先（BigQuery）",
         "シンクで取り込んだログはデータセットに蓄積され、ストレージ課金とクエリ課金が継続的に発生します。",
         "BigQuery ストレージ / クエリ",
         "log_sinks の retention_days（テーブル有効期限）を設定し、不要ログをフィルタで除外する。"),
        ("ログ集約シンク先（Cloud Storage）",
         "バケットへ蓄積されるログ量とストレージクラス・保持期間に応じて課金が継続します。",
         "Cloud Storage 保管 / オペレーション",
         "ライフサイクルルールで古いログを安価なクラスへ移行・自動削除する。retention_days を設定する。"),
        ("Cloud Monitoring（指標取り込み）",
         "監視対象の追加や、ログベース指標・カスタム指標の利用量に応じて取り込み課金が発生します。",
         "Monitoring 指標取り込み",
         "不要なログベース指標・カスタム指標を作りすぎない。監視対象スコープを必要範囲に保つ。"),
        ("ログの長期保持（_Default 等）",
         "ログバケットの保持日数を既定より延長すると、保持量に応じた課金が発生します。",
         "Cloud Logging 保管",
         "要件に応じた保持日数に設定する。長期保管は安価な BigQuery / GCS への集約側で行う。"),
        ("ネットワーク下り（外部への通信）",
         "外部への下りトラフィックや Cloud NAT 等のネットワーク機能は従量課金の対象です。",
         "下りトラフィック / Cloud NAT",
         "不要な外部通信を抑制し、リージョン構成・経路を最適化する。"),
    ]
    records = [{
        "区分": n[0], "コスト発生の要因": n[1], "主な課金対象": n[2], "コスト抑制の推奨": n[3],
    } for n in notes]
    row = table(ws, row, ["区分", "コスト発生の要因", "主な課金対象", "コスト抑制の推奨"], records,
                col_keys=["区分", "コスト発生の要因", "主な課金対象", "コスト抑制の推奨"])
    c = ws.cell(row=row, column=1,
                value="※ 上記は一般的な注意事項です。予算アラート（8-3）の設定とあわせて、定期的なコストレビューを推奨します。"
                      "VPC-SC・組織ポリシー・予算アラート自体に追加課金は発生しません。")
    c.font = F_NOTE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)


# ------------------------------------------------------------------------------
# メイン
# ------------------------------------------------------------------------------
def derive_structural_folders():
    """L0 Bootstrap で組織直下に常に作成される基盤標準フォルダ（SSoT には存在しない）。
    terraform/0_bootstrap/main.tf の google_folder.admin / google_folder.network に対応。"""
    return [
        {"resource_name": "admin", "parent_name": "組織直下",
         "purpose": "管理系プロジェクト（logsink / monitoring）の配置先"},
        {"resource_name": "network", "parent_name": "組織直下",
         "purpose": "ネットワーク基盤（Shared VPC ホスト）の配置先"},
    ]


def derive_mgmt_projects(prefix, tfvars):
    """基盤標準で作成される管理プロジェクト（SSoT 外）。
    1_core/base/{logsink,monitoring,vpc-host} に対応し、配置先フォルダも併記する。
    返り値 [(用途, プロジェクトID, 配置先フォルダ, 説明), ...]。"""
    prefix = prefix or "<prefix>"
    items = [
        ("中央ログ集約", f"{prefix}-logsink", "admin", "監査ログ等を集約・保管する管理プロジェクト"),
        ("中央モニタリング", f"{prefix}-monitoring", "admin", "各プロジェクトを監視対象とするモニタリング管理プロジェクト"),
    ]
    if str(tfvars.get("enable_vpc_host_projects", "false")).lower() == "true":
        items.append(("共有VPCホスト(本番)", f"{prefix}-vpc-prod", "network", "本番系 Shared VPC ホストプロジェクト"))
        items.append(("共有VPCホスト(開発)", f"{prefix}-vpc-dev", "network", "開発系 Shared VPC ホストプロジェクト"))
    return items


def main():
    today = datetime.date.today().strftime("%Y-%m-%d")
    stamp = datetime.date.today().strftime("%Y%m%d")

    tfvars = read_tfvars(COMMON_TFVARS)
    domain = read_domain(DOMAIN_ENV)
    org_domain = tfvars.get("organization_domain") or domain or "（未設定）"

    meta = {
        "customer": os.environ.get("DELIVERY_CUSTOMER", "（顧客名を記入）"),
        "vendor": os.environ.get("DELIVERY_VENDOR", "株式会社イー・エージェンシー"),
        "author": os.environ.get("DELIVERY_AUTHOR", "（作成者を記入）"),
        "version": os.environ.get("DELIVERY_VERSION", "1.0"),
        "doc_no": os.environ.get("DELIVERY_DOCNO", f"GCP-FND-{stamp}"),
        "date": today,
        "org_domain": org_domain,
        "org_id": tfvars.get("organization_id", "（未設定）"),
    }

    if not os.path.exists(XLSX_PATH):
        print(f"⚠️  SSoT '{XLSX_PATH}' が見つかりません。`make template`→`make generate` 後に実行してください。")
        return

    wb_src = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    _, resources = read_sheet(wb_src, "resources")
    _, org_policies = read_sheet(wb_src, "org_policies")
    _, log_sinks = read_sheet(wb_src, "log_sinks")
    _, alert_defs = read_sheet(wb_src, "alert_definitions")
    _, notifications = read_sheet(wb_src, "notifications")
    _, subnets = read_sheet(wb_src, "shared_vpc_subnets")
    _, perimeters = read_sheet(wb_src, "vpc_sc_perimeters")
    _, tag_definitions = read_sheet(wb_src, "tag_definitions")

    folders = [r for r in resources if str(r.get("resource_type")).strip().lower() == "folder"]
    projects = [r for r in resources if str(r.get("resource_type")).strip().lower() == "project"]
    adopted = [p for p in projects if str(p.get("existing_project_id") or "").strip()]
    budget_projects = [p for p in projects if str(p.get("budget_amount") or "").strip()]
    monitoring_targets = [p for p in projects if str(p.get("central_monitoring")).lower() == "true"]

    ctx = {
        "tfvars": tfvars,
        "resources": resources,
        "structural_folders": derive_structural_folders(),
        "folders": folders,
        "projects": projects,
        "adopted_projects": adopted,
        "budget_projects": budget_projects,
        "monitoring_targets": monitoring_targets,
        "org_policies": org_policies,
        "log_sinks": log_sinks,
        "alert_defs": alert_defs,
        "notifications": notifications,
        "subnets": subnets,
        "perimeters": perimeters,
        "tag_definitions": tag_definitions,
        "org_domain": meta["org_domain"],
        "mgmt_projects": derive_mgmt_projects(tfvars.get("project_id_prefix"), tfvars),
    }

    wb = Workbook()
    wb.remove(wb.active)
    build_cover(wb, meta)
    build_revision(wb, meta)
    build_toc(wb)
    build_summary(wb, meta, ctx)
    build_overview(wb, meta, ctx)
    build_folders(wb, ctx)
    build_projects(wb, ctx)
    build_org_policies(wb, ctx)
    build_log_sinks(wb, ctx)
    build_bigquery(wb, ctx)
    build_monitoring(wb, ctx)
    build_network(wb, ctx)
    build_tags_labels(wb, ctx)
    build_groups_iam(wb, ctx)
    build_cost_notes(wb, ctx)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, f"GCP基盤構築_設定明細書_{stamp}.xlsx")
    wb.save(out_path)
    rel = os.path.relpath(out_path, ROOT_DIR)
    print(f"✅ 納品物（構築設定明細書）を生成しました: {rel}")
    print(f"   - 対象組織: {meta['org_domain']}  プロジェクト数: {len(projects)}  フォルダ数: {len(folders)}")
    print("   ℹ️  表紙のメタ情報は環境変数で上書きできます: "
          "DELIVERY_CUSTOMER / DELIVERY_AUTHOR / DELIVERY_VERSION / DELIVERY_DOCNO")


if __name__ == "__main__":
    main()
